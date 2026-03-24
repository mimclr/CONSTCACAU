import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import threading
import os
import sys

# Tentar importar dependências
try:
    import openpyxl
    import pandas as pd
    DEPS_OK = True
except ImportError:
    DEPS_OK = False

# ─── Paleta de cores Labor Rural ────────────────────────────────────────────
COLORS = {
    "bg_dark":     "#0D1F13",
    "bg_card":     "#152A1A",
    "bg_input":    "#1A3320",
    "green_main":  "#2D7A45",
    "green_light": "#3DAF60",
    "green_accent":"#56D97F",
    "text_white":  "#F0F4F1",
    "text_muted":  "#8BA898",
    "text_label":  "#C2D9C8",
    "warn_red":    "#E05555",
    "warn_yellow": "#E0B455",
    "warn_green":  "#56D97F",
    "border":      "#253D2C",
    "scrollbar":   "#2D5038",
}

FONT_TITLE  = ("Helvetica", 22, "bold")
FONT_SUB    = ("Helvetica", 13, "bold")
FONT_BODY   = ("Helvetica", 11)
FONT_SMALL  = ("Helvetica", 9)
FONT_BADGE  = ("Helvetica", 10, "bold")
FONT_MONO   = ("Courier", 10)

# ─── Lógica de análise ──────────────────────────────────────────────────────

def analisar_planilha(filepath):
    """Roda todas as verificações e retorna lista de inconsistências."""
    issues = []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return [{"aba": "ERRO", "linha": "-", "tipo": "CRÍTICO",
                 "descricao": f"Não foi possível abrir o arquivo: {e}"}]

    abas = wb.sheetnames

    # ── helpers ──
    def to_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def normalizar(s):
        if s is None:
            return ""
        return str(s).strip().upper()

    def eh_mao_de_obra(elemento):
        el = normalizar(elemento)
        palavras_mdo = ["MÃO DE OBRA", "MAO DE OBRA", "M.O.", "M.O", "MO ", "DIARISTA",
                        "DIÁRIA", "DIARIA", "MEEIRO", "CONTRATADA", "FUNCIONÁRIO",
                        "FUNCIONARIO", "TRABALHADOR"]
        return any(p in el for p in palavras_mdo)

    # ────────────────────────────────────────────────────────────────────────
    # 1. ABA TALHAO – coletar talhões em produção
    # ────────────────────────────────────────────────────────────────────────
    talhoes_producao = set()
    talhoes_todos = set()

    if "TALHAO" in abas:
        ws = wb["TALHAO"]
        rows = list(ws.iter_rows(values_only=True))
        header = [normalizar(c) for c in rows[0]] if rows else []

        col_nome   = header.index("TALHAO")   if "TALHAO"   in header else None
        col_estagio = header.index("ESTÁGIO") if "ESTÁGIO"  in header else (
                      header.index("ESTAGIO") if "ESTAGIO"  in header else None)

        for i, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            nome = normalizar(row[col_nome]) if col_nome is not None else ""
            estagio = normalizar(row[col_estagio]) if col_estagio is not None else ""
            if nome:
                talhoes_todos.add(nome)
                if estagio == "PRODUÇÃO" or estagio == "PRODUCAO":
                    talhoes_producao.add(nome)
    else:
        issues.append({"aba": "TALHAO", "linha": "-", "tipo": "CRÍTICO",
                       "descricao": "Aba TALHAO não encontrada na planilha."})

    # ────────────────────────────────────────────────────────────────────────
    # 2. ABA INVENTARIO
    # ────────────────────────────────────────────────────────────────────────
    if "INVENTARIO" in abas:
        ws = wb["INVENTARIO"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [normalizar(c) for c in rows[0]]

            def ci(nome):
                return header.index(nome) if nome in header else None

            c_desc    = ci("DESCRIÇÃO") or ci("DESCRICAO") or 4
            c_fab     = ci("DATA DE FABRICAÇÃO") or ci("DATA DE FABRICACAO") or 5
            c_aquisic = ci("DATA DE AQUISIÇÃO") or ci("DATA DE AQUISICAO") or 6
            c_vpago   = ci("VALOR PAGO (R$)") or 8
            c_vnovo   = ci("VALOR DO ITEM NOVO (R$)") or 7

            for i, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue

                desc   = row[c_desc] if c_desc < len(row) else None
                fab    = row[c_fab]  if c_fab  < len(row) else None
                acq    = row[c_aquisic] if c_aquisic < len(row) else None
                vpago  = to_float(row[c_vpago])  if c_vpago  < len(row) else None
                vnovo  = to_float(row[c_vnovo])  if c_vnovo  < len(row) else None

                desc_str = str(desc) if desc else f"Linha {i}"

                # Verificar valor pago
                for label, val in [("Valor Pago", vpago), ("Valor Novo", vnovo)]:
                    if val is not None:
                        if val < 100:
                            issues.append({
                                "aba": "INVENTARIO", "linha": i, "tipo": "ALERTA",
                                "descricao": f"{desc_str} | {label} R$ {val:,.2f} está abaixo de R$ 100,00 — verificar se está correto."
                            })
                        elif val > 500000:
                            issues.append({
                                "aba": "INVENTARIO", "linha": i, "tipo": "ALERTA",
                                "descricao": f"{desc_str} | {label} R$ {val:,.2f} está acima de R$ 500.000,00 — confirmar o valor."
                            })

                # Data fabricação > data aquisição
                if fab and acq:
                    try:
                        import datetime
                        fab_dt = fab if isinstance(fab, datetime.datetime) else None
                        acq_dt = acq if isinstance(acq, datetime.datetime) else None
                        if fab_dt and acq_dt and fab_dt > acq_dt:
                            issues.append({
                                "aba": "INVENTARIO", "linha": i, "tipo": "ERRO",
                                "descricao": f"{desc_str} | Data de fabricação ({fab_dt.strftime('%d/%m/%Y')}) é posterior à data de aquisição ({acq_dt.strftime('%d/%m/%Y')})."
                            })
                    except Exception:
                        pass
    else:
        issues.append({"aba": "INVENTARIO", "linha": "-", "tipo": "CRÍTICO",
                       "descricao": "Aba INVENTARIO não encontrada na planilha."})

    # ────────────────────────────────────────────────────────────────────────
    # 3. ABA PRODUCAO
    # ────────────────────────────────────────────────────────────────────────
    if "PRODUCAO" in abas:
        ws = wb["PRODUCAO"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [normalizar(c) for c in rows[0]]

            def cp(nome):
                for h in header:
                    if nome in h:
                        return header.index(h)
                return None

            c_talhao  = cp("TALHÃO") or cp("TALHAO") or 5
            c_rateio  = cp("RATEIO") or 6
            c_ptotal  = cp("PRODUÇÃO TOTAL") or cp("PRODUCAO TOTAL") or 7
            c_ptalhao = cp("PRODUÇÃO TALHAO") or cp("PRODUCAO TALHAO") or 8
            c_mes     = cp("MÊS") or cp("MES") or 4

            data_rows = [r for r in rows[1:] if any(r)]

            # Agrupar por mês + identificador de lançamento rateado
            from collections import defaultdict
            grupos_rateio = defaultdict(list)

            for i, row in enumerate(data_rows, start=2):
                talhao = normalizar(row[c_talhao]) if c_talhao < len(row) else ""
                rateio = normalizar(row[c_rateio]) if c_rateio < len(row) else ""
                ptotal = to_float(row[c_ptotal])   if c_ptotal < len(row) else None
                mes    = row[c_mes] if c_mes < len(row) else None

                if rateio == "SIM" and mes:
                    mes_key = str(mes)[:7]
                    grupos_rateio[(mes_key, ptotal)].append({
                        "linha": i, "talhao": talhao, "ptalhao": to_float(row[c_ptalhao]) if c_ptalhao < len(row) else None
                    })

            # Verificar rateio: mesmos talhões de produção e mesmo valor
            for (mes_key, ptotal), grupo in grupos_rateio.items():
                talhoes_no_grupo = {g["talhao"] for g in grupo}
                # Checar se todos os talhões de produção estão no grupo
                faltando = talhoes_producao - talhoes_no_grupo
                if faltando:
                    issues.append({
                        "aba": "PRODUCAO", "linha": grupo[0]["linha"], "tipo": "ERRO",
                        "descricao": f"Rateio em {mes_key} | Talhões de produção faltando no lançamento: {', '.join(sorted(faltando))}."
                    })
                # Checar se o valor total é igual em todos os talhões
                valores = [g["ptalhao"] for g in grupo if g["ptalhao"] is not None]
                if valores and len(set(round(v, 4) for v in valores)) > 1:
                    issues.append({
                        "aba": "PRODUCAO", "linha": grupo[0]["linha"], "tipo": "ERRO",
                        "descricao": f"Rateio em {mes_key} | Valores de produção por talhão divergentes: {[round(v,2) for v in valores]}. Com rateio, devem ser iguais."
                    })
    else:
        issues.append({"aba": "PRODUCAO", "linha": "-", "tipo": "CRÍTICO",
                       "descricao": "Aba PRODUCAO não encontrada na planilha."})

    # ────────────────────────────────────────────────────────────────────────
    # 4. ABA DESPESAS
    # ────────────────────────────────────────────────────────────────────────
    ATIVIDADES_COM_MO = {
        "CONDUÇÃO DE LAVOURA", "CONDUCAO DE LAVOURA",
        "ADUBAÇÃO VIA SOLO", "ADUBACAO VIA SOLO",
        "ADUBAÇÃO VIA FOLHA", "ADUBACAO VIA FOLHA",
        "CONTROLE DE PRAGAS E DOENÇAS", "CONTROLE DE PRAGAS E DOENCAS",
        "CONTROLE DE PLANTAS DANINHAS",
        "COLHEITA",
        "PÓS-COLHEITA", "POS-COLHEITA", "PÓS COLHEITA", "POS COLHEITA",
    }

    if "DESPESAS" in abas:
        ws = wb["DESPESAS"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [normalizar(c) for c in rows[0]]

            def cd(nome):
                for h in header:
                    if nome in h:
                        return header.index(h)
                return None

            c_talhao  = cd("TALHÃO")   or cd("TALHAO") or 4
            c_mes     = cd("MÊS")      or cd("MES") or 5
            c_ativ    = cd("ATIVIDADE")or 6
            c_elem    = cd("ELEMENTO") or 7
            c_rateio  = cd("RATEIO")   or 8
            c_qtd_t   = cd("QUANTIDADE TOTAL") or 9
            c_qtd_ta  = cd("QUANTIDADE TALHÃO") or cd("QUANTIDADE TALHAO") or 10
            c_unidade = cd("UNIDADE")  or 11
            c_vunit   = cd("VALOR UNITÁRIO") or cd("VALOR UNITARIO") or 12
            c_area    = cd("ÁREA")     or cd("AREA") or 13
            c_vtotal  = cd("VALOR TOTAL (R$)") or 14
            c_vrsha   = cd("VALOR TOTAL (R$/HA)") or cd("R$/HA") or 15

            data_rows_d = [r for r in rows[1:] if any(r)]

            # Estrutura para checar recorrência administrativa
            from collections import defaultdict
            admin_por_mes_elem = defaultdict(list)   # (mes, elemento) -> [linhas]
            despesas_por_rateio = defaultdict(list)  # (mes, atividade, elemento) -> [linhas]
            despesas_iguais_check = []               # lista de (mes, ativ, elem, vtotal, linha)

            for i, row in enumerate(data_rows_d, start=2):
                def safe(col):
                    try:
                        return row[col] if col is not None and col < len(row) else None
                    except IndexError:
                        return None

                talhao  = normalizar(safe(c_talhao))
                mes     = safe(c_mes)
                ativ    = normalizar(safe(c_ativ))
                elem    = normalizar(safe(c_elem))
                rateio  = normalizar(safe(c_rateio))
                unidade = normalizar(safe(c_unidade))
                vunit   = to_float(safe(c_vunit))
                vtotal  = to_float(safe(c_vtotal))
                vrsha   = to_float(safe(c_vrsha))
                mes_key = str(mes)[:7] if mes else "?"

                # ── 4.1 Atividades que exigem mão de obra ──
                ativ_norm = ativ.upper()
                requer_mo = any(a in ativ_norm for a in ATIVIDADES_COM_MO)
                if requer_mo and not eh_mao_de_obra(elem):
                    issues.append({
                        "aba": "DESPESAS", "linha": i, "tipo": "ALERTA",
                        "descricao": f"[{ativ}] Elemento '{elem}' não é mão de obra. Verificar se há lançamento de M.O. correspondente neste mês ({mes_key}) para o talhão '{talhao}'."
                    })

                # ── 4.2 Manutenção de máquinas lançada como Administração ──
                if "MANUTENÇÃO" in elem and "MÁQUINAS" in elem and ativ == "ADMINISTRAÇÃO":
                    issues.append({
                        "aba": "DESPESAS", "linha": i, "tipo": "ERRO",
                        "descricao": f"Manutenção de Máquinas/Equipamentos lançada como ADMINISTRAÇÃO (talhão: {talhao}, mês: {mes_key}). Deve ser lançada na atividade correta."
                    })

                # ── 4.3 R$/ha acima de 5.000 ──
                if vrsha is not None and vrsha > 5000:
                    issues.append({
                        "aba": "DESPESAS", "linha": i, "tipo": "ALERTA",
                        "descricao": f"[{ativ}] '{elem}' | R$/ha = R$ {vrsha:,.2f} ultrapassa R$ 5.000/ha — confirmar valor (talhão: {talhao}, mês: {mes_key})."
                    })

                # ── 4.4 Valor unitário acima de 5.000 ──
                if vunit is not None and vunit > 5000:
                    issues.append({
                        "aba": "DESPESAS", "linha": i, "tipo": "ALERTA",
                        "descricao": f"[{ativ}] '{elem}' | Valor unitário R$ {vunit:,.2f} acima de R$ 5.000,00 — confirmar (talhão: {talhao}, mês: {mes_key})."
                    })

                # ── 4.5 Valor unitário abaixo de R$1,00 ──
                if vunit is not None and vunit < 1:
                    issues.append({
                        "aba": "DESPESAS", "linha": i, "tipo": "ALERTA",
                        "descricao": f"[{ativ}] '{elem}' | Valor unitário R$ {vunit:,.4f} abaixo de R$ 1,00 — possível erro de digitação (talhão: {talhao}, mês: {mes_key})."
                    })

                # ── 4.6 Kg ou Litros com unitário acima de R$200 ──
                if unidade in ("KG", "LITROS", "L", "LITRO") and vunit is not None and vunit > 200:
                    issues.append({
                        "aba": "DESPESAS", "linha": i, "tipo": "ALERTA",
                        "descricao": f"[{ativ}] '{elem}' | Unidade {unidade} com valor unitário R$ {vunit:,.2f} acima de R$ 200,00 — confirmar (talhão: {talhao}, mês: {mes_key})."
                    })

                # ── 4.7 Coletar para verificar rateio ──
                if rateio == "SIM":
                    despesas_por_rateio[(mes_key, ativ, elem, round(vtotal or 0, 2) if vtotal else None)].append({
                        "linha": i, "talhao": talhao, "vtotal": vtotal
                    })

                # ── 4.8 Coletar para checar despesas iguais ──
                despesas_iguais_check.append((mes_key, ativ, elem, round(vtotal or 0, 4), talhao, i))

                # ── 4.9 Administração: coletar para checar recorrência ──
                if ativ == "ADMINISTRAÇÃO":
                    admin_por_mes_elem[(mes_key, elem)].append(i)

            # ── 4.10 Rateio em despesas: todos talhões presentes + valor igual ──
            for (mes_key, ativ, elem, vtotal_ref), grupo in despesas_por_rateio.items():
                talhoes_no_grupo = {g["talhao"] for g in grupo}
                faltando = talhoes_todos - talhoes_no_grupo
                if faltando:
                    issues.append({
                        "aba": "DESPESAS", "linha": grupo[0]["linha"], "tipo": "ERRO",
                        "descricao": f"[{ativ}] '{elem}' com rateio em {mes_key} | Talhões faltando: {', '.join(sorted(faltando))}."
                    })
                # Verificar se valor total é igual em todos
                valores = [g["vtotal"] for g in grupo if g["vtotal"] is not None]
                if valores:
                    vmin, vmax = min(valores), max(valores)
                    # Tolerar diferença de até 0.01 (arredondamento)
                    if abs(vmax - vmin) > 0.05:
                        issues.append({
                            "aba": "DESPESAS", "linha": grupo[0]["linha"], "tipo": "ERRO",
                            "descricao": f"[{ativ}] '{elem}' com rateio em {mes_key} | Valores totais por talhão divergentes (min: R$ {vmin:,.2f} / max: R$ {vmax:,.2f}). Com rateio devem ser iguais."
                        })

            # ── 4.11 Despesas com valores exatamente iguais no mesmo mês ──
            from collections import Counter
            chave_igual = Counter()
            chave_linhas = defaultdict(list)
            for (mes_key, ativ, elem, vtotal, talhao, linha) in despesas_iguais_check:
                k = (mes_key, ativ, elem, vtotal, talhao)
                chave_igual[k] += 1
                chave_linhas[k].append(linha)

            for k, cnt in chave_igual.items():
                if cnt > 1:
                    mes_key, ativ, elem, vtotal, talhao = k
                    linhas_str = ", ".join(str(l) for l in chave_linhas[k])
                    issues.append({
                        "aba": "DESPESAS", "linha": chave_linhas[k][0], "tipo": "ALERTA",
                        "descricao": f"[{ativ}] '{elem}' | Lançamentos com valor R$ {vtotal:,.2f} idênticos no mesmo mês {mes_key} para o talhão '{talhao}' (linhas: {linhas_str}) — possível duplicidade."
                    })

            # ── 4.12 Recorrência administrativa: mesmo elemento lançado em todos os meses ──
            meses_com_admin = set(k[0] for k in admin_por_mes_elem.keys())
            elementos_admin = set(k[1] for k in admin_por_mes_elem.keys())
            for elem in elementos_admin:
                meses_com_elem = {k[0] for k in admin_por_mes_elem.keys() if k[1] == elem}
                if len(meses_com_elem) >= 2:
                    # Checar se há meses sem esse elemento quando outros têm
                    meses_faltando = meses_com_admin - meses_com_elem
                    if meses_faltando:
                        issues.append({
                            "aba": "DESPESAS", "linha": "-", "tipo": "INFO",
                            "descricao": f"Administração | '{elem}' está lançado em {sorted(meses_com_elem)} mas falta nos meses: {sorted(meses_faltando)} — verificar recorrência."
                        })
    else:
        issues.append({"aba": "DESPESAS", "linha": "-", "tipo": "CRÍTICO",
                       "descricao": "Aba DESPESAS não encontrada na planilha."})

    # ────────────────────────────────────────────────────────────────────────
    # 5. ABA VENDAS
    # ────────────────────────────────────────────────────────────────────────
    if "VENDAS" in abas:
        ws = wb["VENDAS"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [normalizar(c) for c in rows[0]]

            def cv(nome):
                for h in header:
                    if nome in h:
                        return header.index(h)
                return None

            c_talhao = cv("TALHÃO") or cv("TALHAO") or 5
            c_mes    = cv("MÊS")    or cv("MES") or 4
            c_preco  = cv("PREÇO")  or cv("PRECO") or 7

            for i, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue
                talhao = normalizar(row[c_talhao]) if c_talhao < len(row) else ""
                mes    = row[c_mes] if c_mes < len(row) else None
                preco  = to_float(row[c_preco]) if c_preco < len(row) else None
                mes_key = str(mes)[:7] if mes else "?"

                if preco is not None and preco > 100:
                    issues.append({
                        "aba": "VENDAS", "linha": i, "tipo": "ALERTA",
                        "descricao": f"Talhão '{talhao}' | Preço de venda R$ {preco:,.2f}/sc acima de R$ 100,00 — confirmar valor (mês: {mes_key})."
                    })
    else:
        issues.append({"aba": "VENDAS", "linha": "-", "tipo": "CRÍTICO",
                       "descricao": "Aba VENDAS não encontrada na planilha."})

    return issues


# ─── Interface gráfica ──────────────────────────────────────────────────────

class LaborRuralApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Análise de Consistência — Labor Rural")
        self.geometry("1100x720")
        self.minsize(900, 600)
        self.configure(bg=COLORS["bg_dark"])
        self._filepath = None
        self._issues = []
        self._build_ui()
        # Centralizar janela
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - 1100) // 2
        y = (sh - 720) // 2
        self.geometry(f"1100x720+{x}+{y}")

    def _build_ui(self):
        # ── Cabeçalho ──
        header = tk.Frame(self, bg=COLORS["bg_card"], pady=0)
        header.pack(fill="x", side="top")

        # Linha verde topo
        accent_bar = tk.Frame(header, bg=COLORS["green_accent"], height=3)
        accent_bar.pack(fill="x", side="top")

        header_inner = tk.Frame(header, bg=COLORS["bg_card"], padx=30, pady=18)
        header_inner.pack(fill="x")

        # Logo / título
        logo_frame = tk.Frame(header_inner, bg=COLORS["bg_card"])
        logo_frame.pack(side="left")

        tk.Label(logo_frame, text="⬡", font=("Helvetica", 28), fg=COLORS["green_accent"],
                 bg=COLORS["bg_card"]).pack(side="left", padx=(0, 10))

        title_frame = tk.Frame(logo_frame, bg=COLORS["bg_card"])
        title_frame.pack(side="left")

        tk.Label(title_frame, text="Análise de Consistência",
                 font=FONT_TITLE, fg=COLORS["text_white"], bg=COLORS["bg_card"]).pack(anchor="w")
        tk.Label(title_frame, text="Labor Rural — Plataforma MIMC",
                 font=FONT_SMALL, fg=COLORS["text_muted"], bg=COLORS["bg_card"]).pack(anchor="w")

        # Badges de contagem (lado direito do header)
        self.badge_frame = tk.Frame(header_inner, bg=COLORS["bg_card"])
        self.badge_frame.pack(side="right", padx=10)
        self._build_badges()

        # ── Área de importação ──
        import_frame = tk.Frame(self, bg=COLORS["bg_dark"], padx=30, pady=16)
        import_frame.pack(fill="x")

        file_card = tk.Frame(import_frame, bg=COLORS["bg_input"],
                             relief="flat", bd=0)
        file_card.pack(fill="x")

        # Borda esquerda colorida
        border_left = tk.Frame(file_card, bg=COLORS["green_main"], width=4)
        border_left.pack(side="left", fill="y")

        file_inner = tk.Frame(file_card, bg=COLORS["bg_input"], padx=16, pady=14)
        file_inner.pack(side="left", fill="x", expand=True)

        tk.Label(file_inner, text="PLANILHA SELECIONADA",
                 font=FONT_SMALL, fg=COLORS["text_muted"], bg=COLORS["bg_input"]).pack(anchor="w")

        self.lbl_file = tk.Label(file_inner, text="Nenhum arquivo selecionado",
                                  font=FONT_BODY, fg=COLORS["text_label"], bg=COLORS["bg_input"])
        self.lbl_file.pack(anchor="w")

        btn_area = tk.Frame(file_card, bg=COLORS["bg_input"], padx=16, pady=14)
        btn_area.pack(side="right")

        self.btn_import = self._make_btn(btn_area, "  Importar Planilha  ", self._on_import,
                                          COLORS["green_main"], COLORS["text_white"])
        self.btn_import.pack(side="left", padx=6)

        self.btn_analisar = self._make_btn(btn_area, "  ▶  Analisar  ", self._on_analisar,
                                            COLORS["green_accent"], COLORS["bg_dark"], state="disabled")
        self.btn_analisar.pack(side="left", padx=6)

        # ── Filtros ──
        filter_bar = tk.Frame(self, bg=COLORS["bg_dark"], padx=30, pady=4)
        filter_bar.pack(fill="x")

        tk.Label(filter_bar, text="FILTRAR POR:", font=FONT_SMALL,
                 fg=COLORS["text_muted"], bg=COLORS["bg_dark"]).pack(side="left", padx=(0, 8))

        self._filter_var = tk.StringVar(value="TODOS")
        filtros = ["TODOS", "ERRO", "ALERTA", "INFO", "CRÍTICO"]
        for f in filtros:
            cor = self._cor_tipo(f)
            rb = tk.Radiobutton(filter_bar, text=f, variable=self._filter_var,
                                value=f, command=self._aplicar_filtro,
                                font=FONT_SMALL, fg=cor, selectcolor=COLORS["bg_dark"],
                                bg=COLORS["bg_dark"], activebackground=COLORS["bg_dark"],
                                activeforeground=cor, bd=0, highlightthickness=0)
            rb.pack(side="left", padx=6)

        self._filter_aba = tk.StringVar(value="TODAS")
        self.filter_aba_menu = ttk.Combobox(filter_bar, textvariable=self._filter_aba,
                                             state="readonly", width=18,
                                             font=FONT_SMALL)
        self.filter_aba_menu["values"] = ["TODAS"]
        self.filter_aba_menu.pack(side="left", padx=10)
        self.filter_aba_menu.bind("<<ComboboxSelected>>", lambda e: self._aplicar_filtro())

        # Estilo do Combobox
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TCombobox",
                         fieldbackground=COLORS["bg_input"],
                         background=COLORS["bg_input"],
                         foreground=COLORS["text_white"],
                         arrowcolor=COLORS["green_accent"])

        # ── Tabela de resultados ──
        table_frame = tk.Frame(self, bg=COLORS["bg_dark"], padx=30, pady=8)
        table_frame.pack(fill="both", expand=True)

        # Header da tabela
        th = tk.Frame(table_frame, bg=COLORS["bg_card"], pady=8)
        th.pack(fill="x")

        for col, w, anchor in [("ABA", 100, "w"), ("LINHA", 55, "center"),
                                ("TIPO", 80, "center"), ("DESCRIÇÃO", 700, "w")]:
            tk.Label(th, text=col, font=FONT_BADGE, fg=COLORS["text_muted"],
                     bg=COLORS["bg_card"], width=w//8, anchor=anchor,
                     padx=10).pack(side="left")

        # Área scrollável
        scroll_container = tk.Frame(table_frame, bg=COLORS["bg_dark"])
        scroll_container.pack(fill="both", expand=True, pady=(2, 0))

        scrollbar = tk.Scrollbar(scroll_container, orient="vertical",
                                  bg=COLORS["scrollbar"], troughcolor=COLORS["bg_dark"],
                                  width=8)
        scrollbar.pack(side="right", fill="y")

        self.canvas = tk.Canvas(scroll_container, bg=COLORS["bg_dark"],
                                 yscrollcommand=scrollbar.set,
                                 highlightthickness=0)
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.canvas.yview)

        self.inner_frame = tk.Frame(self.canvas, bg=COLORS["bg_dark"])
        self.canvas_window = self.canvas.create_window((0, 0), window=self.inner_frame, anchor="nw")

        self.inner_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        # Estado vazio
        self.lbl_empty = tk.Label(self.inner_frame,
                                   text="Importe uma planilha e clique em Analisar para visualizar os resultados.",
                                   font=FONT_BODY, fg=COLORS["text_muted"],
                                   bg=COLORS["bg_dark"], pady=40)
        self.lbl_empty.pack()

        # ── Status bar ──
        self.status_bar = tk.Frame(self, bg=COLORS["bg_card"], height=28)
        self.status_bar.pack(fill="x", side="bottom")

        accent_bar2 = tk.Frame(self.status_bar, bg=COLORS["border"], height=1)
        accent_bar2.pack(fill="x", side="top")

        self.lbl_status = tk.Label(self.status_bar, text="Pronto.",
                                    font=FONT_SMALL, fg=COLORS["text_muted"],
                                    bg=COLORS["bg_card"], padx=16, pady=4)
        self.lbl_status.pack(side="left")

        self.lbl_total = tk.Label(self.status_bar, text="",
                                   font=FONT_SMALL, fg=COLORS["green_accent"],
                                   bg=COLORS["bg_card"], padx=16)
        self.lbl_total.pack(side="right")

    def _make_btn(self, parent, text, cmd, bg, fg, state="normal"):
        btn = tk.Button(parent, text=text, command=cmd,
                         font=FONT_BADGE, bg=bg, fg=fg,
                         relief="flat", bd=0, padx=14, pady=8,
                         cursor="hand2", state=state,
                         activebackground=COLORS["green_light"],
                         activeforeground=COLORS["bg_dark"])
        return btn

    def _build_badges(self):
        for w in self.badge_frame.winfo_children():
            w.destroy()
        if not self._issues:
            return
        contagem = {}
        for issue in self._issues:
            t = issue["tipo"]
            contagem[t] = contagem.get(t, 0) + 1

        for tipo, cnt in contagem.items():
            cor = self._cor_tipo(tipo)
            badge = tk.Frame(self.badge_frame, bg=COLORS["bg_dark"],
                              padx=10, pady=4)
            badge.pack(side="left", padx=4)
            tk.Label(badge, text=f"● {tipo}", font=FONT_SMALL,
                     fg=cor, bg=COLORS["bg_dark"]).pack(side="left")
            tk.Label(badge, text=f"  {cnt}", font=FONT_BADGE,
                     fg=COLORS["text_white"], bg=COLORS["bg_dark"]).pack(side="left")

    def _cor_tipo(self, tipo):
        return {
            "ERRO":    COLORS["warn_red"],
            "CRÍTICO": COLORS["warn_red"],
            "ALERTA":  COLORS["warn_yellow"],
            "INFO":    COLORS["text_muted"],
            "TODOS":   COLORS["text_label"],
        }.get(tipo, COLORS["text_label"])

    def _on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_import(self):
        fp = filedialog.askopenfilename(
            title="Selecionar planilha MIMC",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if fp:
            self._filepath = fp
            nome = os.path.basename(fp)
            self.lbl_file.config(text=nome, fg=COLORS["green_accent"])
            self.btn_analisar.config(state="normal")
            self._set_status(f"Arquivo carregado: {nome}")

    def _on_analisar(self):
        if not self._filepath:
            return
        self.btn_analisar.config(state="disabled", text="  ⏳ Analisando...")
        self._set_status("Analisando planilha...")
        self._limpar_tabela()

        def run():
            issues = analisar_planilha(self._filepath)
            self.after(0, lambda: self._mostrar_resultados(issues))

        threading.Thread(target=run, daemon=True).start()

    def _mostrar_resultados(self, issues):
        self._issues = issues
        self._build_badges()
        self._atualizar_filtros_aba()
        self._aplicar_filtro()
        total = len(issues)
        self.btn_analisar.config(state="normal", text="  ▶  Analisar  ")
        self.lbl_total.config(text=f"Total: {total} ocorrências")
        self._set_status(f"Análise concluída. {total} inconsistências encontradas.")

    def _atualizar_filtros_aba(self):
        abas = ["TODAS"] + sorted(set(i["aba"] for i in self._issues))
        self.filter_aba_menu["values"] = abas
        self._filter_aba.set("TODAS")

    def _aplicar_filtro(self):
        tipo_f = self._filter_var.get()
        aba_f  = self._filter_aba.get()
        filtrados = self._issues
        if tipo_f != "TODOS":
            filtrados = [i for i in filtrados if i["tipo"] == tipo_f]
        if aba_f != "TODAS":
            filtrados = [i for i in filtrados if i["aba"] == aba_f]
        self._renderizar_tabela(filtrados)

    def _limpar_tabela(self):
        for w in self.inner_frame.winfo_children():
            w.destroy()

    def _renderizar_tabela(self, issues):
        self._limpar_tabela()
        if not issues:
            tk.Label(self.inner_frame,
                     text="Nenhuma inconsistência encontrada com os filtros selecionados." if self._issues else
                          "Importe uma planilha e clique em Analisar.",
                     font=FONT_BODY, fg=COLORS["text_muted"],
                     bg=COLORS["bg_dark"], pady=40).pack()
            return

        for idx, issue in enumerate(issues):
            bg = COLORS["bg_card"] if idx % 2 == 0 else COLORS["bg_dark"]
            row_frame = tk.Frame(self.inner_frame, bg=bg, pady=0)
            row_frame.pack(fill="x")

            cor = self._cor_tipo(issue["tipo"])

            # Barra lateral colorida
            tk.Frame(row_frame, bg=cor, width=3).pack(side="left", fill="y")

            # ABA
            tk.Label(row_frame, text=issue["aba"], font=FONT_SMALL,
                     fg=COLORS["text_muted"], bg=bg, width=12,
                     anchor="w", padx=8, pady=8).pack(side="left")

            # LINHA
            tk.Label(row_frame, text=str(issue["linha"]), font=FONT_SMALL,
                     fg=COLORS["text_muted"], bg=bg, width=6,
                     anchor="center").pack(side="left")

            # TIPO badge
            badge_bg = self._badge_bg(issue["tipo"])
            badge_frame = tk.Frame(row_frame, bg=badge_bg, padx=6, pady=2)
            badge_frame.pack(side="left", padx=8)
            tk.Label(badge_frame, text=issue["tipo"], font=FONT_SMALL,
                     fg=cor, bg=badge_bg).pack()

            # DESCRIÇÃO
            tk.Label(row_frame, text=issue["descricao"], font=FONT_SMALL,
                     fg=COLORS["text_label"], bg=bg,
                     wraplength=700, justify="left",
                     anchor="w", padx=8).pack(side="left", fill="x", expand=True)

    def _badge_bg(self, tipo):
        return {
            "ERRO":    "#3A1A1A",
            "CRÍTICO": "#3A1A1A",
            "ALERTA":  "#3A2A10",
            "INFO":    "#1A2530",
        }.get(tipo, COLORS["bg_input"])

    def _set_status(self, msg):
        self.lbl_status.config(text=msg)


# ─── Entry point ────────────────────────────────────────────────────────────

def main():
    if not DEPS_OK:
        # Tentar instalar dependências
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install",
                               "openpyxl", "pandas", "--quiet"])
        import importlib
        import openpyxl
        import pandas as pd

    app = LaborRuralApp()
    app.mainloop()


if __name__ == "__main__":
    main()
